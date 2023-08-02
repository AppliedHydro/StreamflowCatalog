import openpyxl as px
import matplotlib.pyplot as plt
import numpy as np
import matplotlib as mpl

#import workbook as iterable dataframe object
workbook = px.open('C:/Users/sjsch/Desktop/Kendra/Streamflow_Catalog.xlsx')
wb = workbook.active

OR_cont,OR_misc,OR_unk = [],[],[]
WA_cont,WA_misc,WA_unk = [],[],[]
ID_cont,ID_misc,ID_unk = [],[],[]
counter = 1

#iterator to sort cell values into sampling type
for row in wb.iter_rows(min_row=2,max_row=26871,min_col=21,max_col=21):
    for cell in row:
        counter += 1
        if str(wb.cell(counter,2).value) == 'USGS':
            continue
        var = str(cell.value).lower()
        if var == '' or var == 'none' or var == 'unknown':
            if str(wb.cell(counter,6).value) == 'Idaho':
                ID_unk.append(var)
            elif str(wb.cell(counter,6).value) == 'Oregon':
                OR_unk.append(var)
            elif str(wb.cell(counter,6).value) == 'Washington':
                WA_unk.append(var)
        elif var == 'continuous':
            if str(wb.cell(counter,6).value) == 'Idaho':
                ID_cont.append(var)
            elif str(wb.cell(counter,6).value) == 'Oregon':
                OR_cont.append(var)
            elif str(wb.cell(counter,6).value) == 'Washington':
                WA_cont.append(var)
        elif var == 'seasonal' or var == 'synoptic':
            if str(wb.cell(counter,6).value) == 'Idaho':
                ID_misc.append(var)
            elif str(wb.cell(counter,6).value) == 'Oregon':
                OR_misc.append(var)
            elif str(wb.cell(counter,6).value) == 'Washington':
                WA_misc.append(var)
                
        else:
            print(var)
            pass

#define data variables for graphic
ID_tot = len(ID_cont) + len(ID_misc) + len(ID_unk)
OR_tot = len(OR_cont) + len(OR_misc) + len(OR_unk)
WA_tot = len(WA_cont) + len(WA_misc) + len(WA_unk)

unk_tot = (len(ID_unk),len(OR_unk),len(WA_unk))
misc_tot = (len(ID_misc),len(OR_misc),len(WA_misc))
cont_tot = (len(ID_cont),len(OR_cont),len(WA_cont))

categories = ['Unknown','Miscellaneous','Continuous(%)']
gages_no = ['Idaho : %s' % ID_tot,'Oregon : %s' % OR_tot,'Washington : %s' % WA_tot]
#convert data to percentages
results = { 
    str(ID_tot):[round((len(ID_unk)/ID_tot)*100,1),round((len(ID_misc)/ID_tot)*100,1),round((len(ID_cont)/ID_tot)*100,1)],
    str(OR_tot):[round((len(OR_unk)/OR_tot)*100,1),round((len(OR_misc)/OR_tot)*100,1),round((len(OR_cont)/OR_tot)*100,1)],
    str(WA_tot):[round((len(WA_unk)/WA_tot)*100,1),round((len(WA_misc)/WA_tot)*100,1),round((len(WA_cont)/WA_tot)*100,1)]
    }
results2 = {
    ID_tot:[ID_cont,ID_misc,ID_unk],
    OR_tot:[OR_cont,OR_misc,OR_unk],
    WA_tot:[WA_cont,WA_misc,WA_unk]
    }
labeltotals = (("{:,}".format(ID_tot),"{:,}".format(OR_tot),"{:,}".format(WA_tot))) #format integers into #,###

#chart graphic              
def PlotDat(results, categories):
    ylabels = np.arange(len(categories))
    labels = list(results.keys())
    data = np.array(list(results.values()))
    data_cum = data.cumsum(axis=1)
    category_colors = plt.colormaps['Wistia'](
        np.linspace(0.15, 0.85, data.shape[1]))
    fig, ax = plt.subplots(figsize=(9.2, 5))
    ax.xaxis.set_visible(False)
    ax.set_xlim(0, np.sum(data, axis=1).max())
    for i, (colname, color) in enumerate(zip(categories, category_colors)): #bar generator for 3x3 distribution
        widths = data[:,i]
        starts = data_cum[:,i] - widths
        rects = ax.barh(labels,widths,left=starts,height=0.5,label=colname,color=color)
        ax.yaxis.set_major_formatter(mpl.ticker.StrMethodFormatter('{x:,.0f}')) #formats into #,### integers
        r,g,b,_ = color
        text_color = 'black'
        ax.bar_label(rects,label_type='center',wrap=False,color=text_color,padding=5.5)
    ax.legend(ncol=len(categories), bbox_to_anchor=(0.25,0),
              loc='upper center', fontsize='small')
    ax.legend(ncol=len(categories), bbox_to_anchor=(0.75,0),
              loc='upper center', fontsize='small')
    ax.set_yticks(ylabels)
    ax.set_yticklabels(labeltotals)
    ax.bar_label(rects, gages_no, padding=0.5)
    ax.set_ylabel('quantity')
    return fig, ax
"""
cont1 = [round((len(ID_cont)/ID_tot)*100,1),round((len(OR_cont)/OR_tot)*100,1),round((len(WA_cont)/WA_tot)*100,1)]
misc1 = [round((len(ID_misc)/ID_tot)*100,1),round((len(OR_misc)/OR_tot)*100,1),round((len(WA_misc)/WA_tot)*100,1)]
unk1 = [round((len(ID_unk)/ID_tot)*100,1),round((len(OR_unk)/OR_tot)*100,1),round((len(WA_unk)/WA_tot)*100,1)]
x=[97.9,7.6,6.5]
y=[1.1,92.1,93.5]
z=[1,0.3,0]

fig, ax = plt.subplots()
width = 0.45
a = ax.bar(gages_no, x, width, label='Continuous',color='gold')
b = ax.bar(gages_no, y, width, label='Discrete',bottom=x,color='darkorange')
c = ax.bar(gages_no, z, width, label='Unknown',color='wheat')
ax.bar_label(a, label_type='center', color='black',padding=3)
ax.bar_label(b, label_type='center', color='black',padding=3)
ax.bar_label(c, label_type='center', color='black',padding=3)
ax.set_ylabel('Percentages')
ax.legend(ncol=len(categories),bbox_to_anchor=(0, 1),loc='lower left', fontsize='small')
plt.show()
"""

cont1 = [round((len(ID_cont)/ID_tot)*100,1),round((len(OR_cont)/OR_tot)*100,1),round((len(WA_cont)/WA_tot)*100,1)]
misc1 = [round((len(ID_misc)/ID_tot)*100,1),round((len(OR_misc)/OR_tot)*100,1),round((len(WA_misc)/WA_tot)*100,1)]
unk1 = [round((len(ID_unk)/ID_tot)*100,1),round((len(OR_unk)/OR_tot)*100,1),round((len(WA_unk)/WA_tot)*100,1)]
x=[1294,993,406]
y=[15,11996,5886]
z=[13,36,0]

fig, ax = plt.subplots()
width = 0.45
a = ax.bar(gages_no, x, width, label='Continuous',color='gold')
b = ax.bar(gages_no, y, width, label='Discrete',bottom=x,color='darkorange')
c = ax.bar(gages_no, z, width, label='Unknown',color='wheat')
ax.bar_label(a, label_type='center', color='black',padding=3)
ax.bar_label(b, label_type='center', color='black',padding=3)
ax.bar_label(c, label_type='center', color='black',padding=3)
ax.set_ylabel('gage totals')
ax.legend(ncol=len(categories),bbox_to_anchor=(0, 1),loc='lower left', fontsize='small')
plt.show()
            
