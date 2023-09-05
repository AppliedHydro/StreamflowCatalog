#####TableUpdater.py#####

####################################################################################
#Used to create a statistics table for gage totals by state, gage types by state,  #
#active/inactive by state, measurement frequency by state                          #
####################################################################################

#Variables
#Edit (line 15) to local file path of streamflow catalog

import openpyxl as px
import numpy as np
from tabulate import tabulate

print("Importing catalog...")
workbook = px.open('C:/Users/sjsch/Desktop/Kendra/Streamflow_Catalog.xlsx')
wb = workbook.active

#list containers to hold gage totals for: overall, continuous, discrete, unknown
IDt, IDc, IDd, IDu = 0,0,0,0
ORt, ORc, ORd, ORu = 0,0,0,0
WAt, WAc, WAd, WAu = 0,0,0,0

ctr = 1 #used to index values
print(f"There are {wb.max_row} rows in the catalog")
for rows in wb.iter_rows(min_row=2,max_row=wb.max_row,min_col=1,max_col=1):
    for cell in rows:
        ctr += 1
        if str(wb.cell(ctr,6).value) == 'Idaho':
            IDt += 1
            if str(wb.cell(ctr,21).value) == 'continuous':
                IDc += 1
            elif str(wb.cell(ctr,21).value) == 'unknown':
                IDu += 1
            else:
                IDd += 1
        elif str(wb.cell(ctr,6).value) == 'Oregon':
            ORt += 1
            if str(wb.cell(ctr,21).value) == 'continuous':
                ORc += 1
            elif str(wb.cell(ctr,21).value) == 'unknown':
                ORu += 1
            else:
                ORd += 1
        elif str(wb.cell(ctr,6).value) == 'Washington':
            WAt += 1
            if str(wb.cell(ctr,21).value) == 'continuous':
                WAc += 1
            elif str(wb.cell(ctr,21).value) == 'unknown':
                WAu += 1
            else:
                WAd += 1
        else:
            print(ctr)
            pass

tbl = [['Idaho',IDt,IDc,IDd,IDu],['Oregon',ORt,ORc,ORd,ORu],['Washington',WAt,WAc,WAd,WAu]]
Headers1 = ['State','Total','Continuous','Discrete','Unknown']
print(tabulate(tbl,headers=Headers1,tablefmt='fancy_grid'))

instant, fifteenmin, thirtymin = 0,0,0
hourly, daily, annual, unknown = 0,0,0,0
monthly = 0
ctr = 1
for rows in wb.iter_rows(min_row=2,max_row=wb.max_row,min_col=20,max_col=20):
    for cell in rows:
        ctr += 1
        if str(cell.value) == 'hourly':
            hourly += 1
        elif str(cell.value) == 'daily':
            daily += 1
        elif str(cell.value) == '15 minutes':
            fifteenmin += 1
        elif str(cell.value) == '30 minutes':
            thirtymin += 1
        elif str(cell.value) == 'instantaneous':
            instant += 1
        elif str(cell.value) == 'anually':
            annual += 1
        elif str(cell.value) == 'unknown':
            unknown += 1
        elif str(cell.value) == 'monthly':
            monthly += 1
        else:
            pass

tbl2 = [[int(instant),int(fifteenmin),int(thirtymin),int(hourly),int(daily),int(monthly),int(annual),int(unknown)]]
Headers2 = ['instantaneous','15 minutes','30 minutes','hourly','daily','monthly','anually','unknown']
print(tabulate(tbl2,headers=Headers2,tablefmt='fancy_grid'))

contcounter,ID,OR,WA = 0,0,0,0
idact,idinact,idunk = 0,0,0
oract,orinact,orunk = 0,0,0
waact,wainact,waunk = 0,0,0
idcanal,idstream,id_unk = 0,0,0
orcanal,orstream,or_unk = 0,0,0
wacanal,wastream,wa_unk = 0,0,0
ctr = 1

for rows in wb.iter_rows(min_row=2,max_row=wb.max_row,min_col=21,max_col=21):
    for cell in rows:
        ctr += 1
        if str(cell.value) == 'continuous' or str(cell.value) == 'seasonal':
            if str(wb.cell(ctr,6).value) == 'Idaho':
                ID += 1
                if str(wb.cell(ctr,17).value) == 'active':
                    idact += 1
                elif str(wb.cell(ctr,17).value) == 'inactive':
                    idinact += 1
                else:
                    idunk += 1
            elif str(wb.cell(ctr,6).value) == 'Oregon':
                OR += 1
                if str(wb.cell(ctr,17).value) == 'active':
                    oract += 1
                elif str(wb.cell(ctr,17).value) == 'inactive':
                    orinact += 1
                else:
                    orunk += 1
            elif str(wb.cell(ctr,6).value) == 'Washington':
                WA += 1
                if str(wb.cell(ctr,17).value) == 'active':
                    waact += 1
                elif str(wb.cell(ctr,17).value) == 'inactive':
                    wainact += 1
                else:
                    waunk += 1
            else:
                pass
        else:
            pass

ctr = 1
for rows in wb.iter_rows(min_row=2,max_row=wb.max_row,min_col=21,max_col=21):
    for cell in rows:
        ctr += 1
        if str(cell.value) == 'continuous' or str(cell.value) == 'seasonal':
            if str(wb.cell(ctr,6).value) == 'Idaho':
                if str(wb.cell(ctr,9).value) == 'canal':
                    idcanal += 1
                elif str(wb.cell(ctr,9).value) == 'stream':
                    idstream += 1
                else:
                    id_unk += 1
            elif str(wb.cell(ctr,6).value) == 'Oregon':
                if str(wb.cell(ctr,9).value) == 'canal':
                    orcanal += 1
                elif str(wb.cell(ctr,9).value) == 'stream':
                    orstream += 1
                else:
                    or_unk += 1
            elif str(wb.cell(ctr,6).value) == 'Washington':
                if str(wb.cell(ctr,9).value) == 'canal':
                    wacanal += 1
                elif str(wb.cell(ctr,9).value) == 'stream':
                    wastream += 1
                else:
                    wa_unk += 1
            else:
                pass
        else:
            pass

sumact,suminact,sumunk = (idact+oract+waact),(idinact+orinact+wainact),(idunk+orunk+waunk)
sumcanal,sumstream,sum_unk = (idcanal+orcanal+wacanal),(idstream+orstream+wastream),(id_unk+or_unk+wa_unk)
tbl3 = [['ID',ID,idact,idinact,idunk,idcanal,idstream,id_unk],['OR',OR,oract,orinact,orunk,orcanal,orstream,or_unk],
        ['WA',WA,waact,wainact,waunk,wacanal,wastream,wa_unk],['PNW total',(WA+OR+ID),sumact,suminact,sumunk,sumcanal,sumstream,sum_unk]]
Headers3 = ['State','Total','Active','Inactive','Unknown','Canal','Stream','Unknown']
print(tabulate(tbl3,headers=Headers3,tablefmt='fancy_grid'))
        
        
