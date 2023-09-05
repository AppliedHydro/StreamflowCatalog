#####activeinactive_bystate_withUSGS.py#####

############################################################
#Plots gages by active/inactive types, includes USGS gages #
############################################################

import openpyxl as px             #handles catalog as dataframe
import matplotlib.pyplot as plt   #visualizes data
import numpy as np                #mathematical library with built-in calculation tools
import matplotlib as mpl

#Variables
#line 17: edit file path to streamflow catalog on local machine

#import catalog and process into openpyxl dataframe
print("Importing Catalog...")
workbook = px.open('C:/Users/sjsch/Desktop/Kendra/Streamflow_Catalog.xlsx')
wb = workbook.active
OR_inact,OR_act = [],[]
ID_inact,ID_act = [],[]
WA_inact,WA_act = [],[]
ID_unk,OR_unk,WA_unk = [],[],[]

#function to sort by state and active designation
print(f"There are {wb.max_row} in the streamflow catalog")
def ActiveCounter():
    counter = 1
    for row in wb.iter_rows(min_col=17,max_col=17,min_row=2,max_row=wb.max_row):
        for cell in row:
            var = str(cell.value).lower()
            counter += 1
            if var == 'active' and str(wb.cell(counter,6).value) == 'Idaho':
                ID_act.append(var)
            elif var == 'inactive' and str(wb.cell(counter,6).value) == 'Idaho':
                ID_inact.append(var)
            elif var == 'active' and str(wb.cell(counter,6).value) == 'Oregon':
                OR_act.append(var)
            elif var == 'inactive' and str(wb.cell(counter,6).value) == 'Oregon':
                OR_inact.append(var)
            elif var == 'active' and str(wb.cell(counter,6).value) == 'Washington':
                WA_act.append(var)
            elif var == 'inactive' and str(wb.cell(counter,6).value) == 'Washington':
                WA_inact.append(var)
            else:
                if str(wb.cell(counter,6).value) == 'Idaho':
                    ID_unk.append(var)
                elif str(wb.cell(counter,6).value) == 'Oregon':
                    OR_unk.append(var)
                elif str(wb.cell(counter,6).value) == 'Washington':
                    WA_unk.append(var)
ActiveCounter()

#designating graphic x,y parameters
active = [len(ID_act),len(OR_act),len(WA_act)]
inactive = [len(ID_inact),len(OR_inact),len(WA_inact)]
unknown = [len(ID_unk),len(OR_unk),len(WA_unk)]
x = np.arange(len(active))
width = 0.3

#chart graphic
fig, ax = plt.subplots()
labels = ['Idaho','Oregon','Washington']
rects1 = ax.bar(x - width/2.2,active, width, label='active',color='gold')
rects2 = ax.bar(x + width/3, inactive, width, label='inactive',color='orange')
rects3 = ax.bar(x + width/1, unknown, width, label='unknown',color='darkgoldenrod')
ax.yaxis.set_major_formatter(mpl.ticker.StrMethodFormatter('{x:,.0f}')) #formats into #,### integers
ax.set_ylabel('quantity')
ax.set_xticks(x, labels)

#defines integer labels for each bar, pulls directly from graph data
ax.bar_label((rects1), labels=[f'{x:,.0f}' for x in rects1.datavalues], padding=1)
ax.bar_label((rects2), labels=[f'{x:,.0f}' for x in rects2.datavalues], padding=1)
ax.bar_label((rects3), labels=[f'{x:,.0f}' for x in rects3.datavalues], padding=1)
ax.legend(fancybox=True)

fig.tight_layout()
plt.show()
