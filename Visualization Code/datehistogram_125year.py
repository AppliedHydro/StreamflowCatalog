from datetime import datetime
import openpyxl as px           #handles catalog data as dataframe object
import matplotlib.pyplot as plt #visualizes data 
import numpy as np              #collections of calculation tools

#import catalog as dataframe object
workbook = px.open('C:/Users/sjsch/Desktop/Kendra/Streamflow_Catalog.xlsx')
wb = workbook.active

years_start = []
years_end = []
unknown = []
counter = 1

#iterate through date values, discard 'unknowns' and convert rest into datetime object
try:
    for row in wb.iter_rows(min_col=15,max_col=15,min_row=2,max_row=26871):
        for cell in row:
            counter += 1
            date_format = ('%Y-%m-%d %H:%M:%S')
            date_format1 = ('%H:%M:%S')
            var = str(cell.value).lower()
            if str(wb.cell(counter,16).value) == 'current':
                cell.value = datetime.now()
                pass
            elif str(wb.cell(counter,21).value).lower() == 'continuous': #filter by continuous
                if str(wb.cell(counter,16).value).lower() == 'none' or str(wb.cell(counter,16).value).lower() == 'unknown':
                    unknown.append(var)
                    continue
                elif var == 'none' or var == 'unknown':
                    unknown.append(var)
                    continue
                date_s = datetime.strptime(var,date_format) #pulls date from excel string as datetime object
                years_start.append(date_s.year)
                date_e = datetime.strptime(str(wb.cell(counter,16).value),date_format)
                years_end.append(date_e.year)
                continue
            else:
                continue
except: #checks to ensure all values have been handled
    print(counter)
    raise

totals = []
counter = 0
try:        #calculate time total between start and end dates
    for i in range(0,25526):
        var = int(years_end[counter]) - int(years_start[counter])
        totals.append(var)
        counter += 1
except:
    pass

#chart graphic
dist1 = totals
dist2 = range(0,126,10)
fig, axs = plt.subplots(tight_layout=True)
axs.hist(dist1,bins=125,range=(0,125),color='orange')
axs.set_xticks(dist2)
axs.set_ylabel('quantity')
axs.set_xlabel('years of operation')
axs.set_xlim([0,125]) #constrain x-axis limitations for viewability
plt.show()
