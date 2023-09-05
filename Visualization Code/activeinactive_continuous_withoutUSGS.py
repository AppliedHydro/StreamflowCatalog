#####activeinactive_continuous_withoutUSGS.py#####

#############################################################################
#charts gages based on active/inactive, subsetted to only continuous gages  #
#excludes USGS gages                                                        #
#############################################################################

#Variables 
#line 17 : edit file path to streamflow catalog local address

import openpyxl as px             #handles catalog as dataframe
import matplotlib.pyplot as plt   #visualizes data
import numpy as np                #mathematical library with built-in calculation tools
import matplotlib as mpl

#import catalog and process into openpyxl dataframe
print("Importing catalog...")
workbook = px.open('C:/Users/sjsch/Desktop/Kendra/Streamflow_Catalog.xlsx')
wb = workbook.active
OR_inact,OR_act = [],[]
ID_inact,ID_act = [],[]
WA_inact,WA_act = [],[]
ID_unk,OR_unk,WA_unk = [],[],[]

#function to sort by state and active designation
print(f"There are {wb.max_row} gages in the catalog")
def ActiveCounter():
    counter = 1
    for row in wb.iter_rows(min_col=21,max_col=21,min_row=2,max_row=wb.max_row):
        for cell in row:
            var = str(cell.value).lower()
            counter += 1
            if var == 'continuous': #subsets by 'continuous' designation
                if str(wb.cell(counter,6).value) == 'Idaho':
                    if str(wb.cell(counter,17).value) == 'active':
                        ID_act.append(var)
                    elif str(wb.cell(counter,17).value) == 'inactive':
                        ID_inact.append(var)
                    else:
                        ID_unk.append(var)
                elif str(wb.cell(counter,6).value) == 'Oregon':
                    if str(wb.cell(counter,17).value) == 'active':
                        OR_act.append(var)
                    elif str(wb.cell(counter,17).value) == 'inactive':
                        OR_inact.append(var)
                    else:
                        OR_unk.append(var)
                elif str(wb.cell(counter,6).value) == 'Washington':
                    if str(wb.cell(counter,17).value) == 'active':
                        WA_act.append(var)
                    elif str(wb.cell(counter,17).value) == 'inactive':
                        WA_inact.append(var)
                    else:
                        WA_unk.append(var)
                    

ActiveCounter()

#designating graphic x,y parameters
labels = ['Idaho','Oregon','Washington']
active = [len(ID_act),len(OR_act),len(WA_act)]
inactive = [len(ID_inact),len(OR_inact),len(WA_inact)]
unknown = [len(ID_unk),len(OR_unk),len(WA_unk)]
x = np.arange(len(active))
width = 0.3

#chart graphic
fig, ax = plt.subplots()
rects1 = ax.bar(x - width/2.2,active, width, label='active',color='gold')
rects2 = ax.bar(x + width/3, inactive, width, label='inactive',color='orange')
rects3 = ax.bar(x + width/1, unknown, width, label='unknown',color='darkgoldenrod')
ax.yaxis.set_major_formatter(mpl.ticker.StrMethodFormatter('{x:,.0f}')) #formats into #,### integers
ax.set_ylabel('quantity')
ax.set_xticks(x, labels)

#defines integer labels for each bar, pulls directly from graph data
ax.bar_label((rects1), padding=1)
ax.bar_label((rects2), labels=[f'{x:,.0f}' for x in rects2.datavalues], padding=1)
ax.bar_label((rects3), labels=[f'{x:,.0f}' for x in rects3.datavalues], padding=1)
ax.legend(fancybox=True)

fig.tight_layout()
plt.show()

