import openpyxl as px             #handles catalog as dataframe
import matplotlib.pyplot as plt   #visualizes data
import numpy as np                #mathematical library with built-in calculation tools
import matplotlib as mpl

workbook = px.open('C:/Users/sjsch/Desktop/Kendra/Streamflow_Catalog.xlsx')
wb = workbook.active

ID_unk, ID_canal, ID_weir, ID_stream = [],[],[],[]
OR_unk, OR_canal, OR_weir, OR_stream = [],[],[],[]
WA_unk, WA_canal, WA_weir, WA_stream = [],[],[],[]
I, O, W = 'Idaho','Oregon','Washington'
exceptions = []
counter = 1

#iterator for sorting values into variable containers
for row in wb.iter_rows(min_col=9,max_col=9,min_row=2,max_row=26883):
    for cell in row:
        counter += 1
        var = str(cell.value).lower()
        var = var.rstrip(' ') #standardize values
        if str(wb.cell(counter,1).value) == 'None':
            continue
        elif var == 'none' or var == 'unknown':
            if wb.cell(counter,6).value == I:
                ID_unk.append(var)
            elif wb.cell(counter,6).value == O:
                OR_unk.append(var)
            elif wb.cell(counter,6).value == W:
                WA_unk.append(var)
        elif var == 'canal':
            if wb.cell(counter,6).value == I:
                ID_canal.append(var)
            elif wb.cell(counter,6).value == O:
                OR_canal.append(var)
            elif wb.cell(counter,6).value == W:
                WA_canal.append(var)
        elif var == 'stream':
            if wb.cell(counter,6).value == I:
                ID_stream.append(var)
            elif wb.cell(counter,6).value == O:
                OR_stream.append(var)
            elif wb.cell(counter,6).value == W:
                WA_stream.append(var)
        elif var == 'weir':
            if wb.cell(counter,6).value == I:
                ID_weir.append(var)
            elif wb.cell(counter,6).value == O:
                OR_stream.append(var)
            elif wb.cell(counter,6).value == W:
                WA_stream.append(var)
        else:
            exceptions.append(str(counter)+ var)

labels = ['Idaho','Oregon','Washington']
canal = [len(ID_canal),len(OR_canal),len(WA_canal)]
weir = [len(ID_weir),len(OR_weir),len(WA_weir)]
stream = [len(ID_stream),len(OR_stream),len(WA_stream)]
unknown = [len(ID_unk),len(OR_unk),len(WA_unk)]
x = np.arange(len(labels))
width = 0.3
fig, ax = plt.subplots()
rects1 = ax.bar(x + width/1, canal, width, label='canal',color='gold')
rects2 = ax.bar(x + width/2.2, weir, width, label='weir', color='lightyellow')
rects3 = ax.bar(x, stream, width, label='stream',color='orange')
rects4 = ax.bar(x - width/1, unknown, width, label='unknown',color='darkgoldenrod')
ax.yaxis.set_major_formatter(mpl.ticker.StrMethodFormatter('{x:,.0f}'))
ax.set_ylabel('quantity')
ax.set_xticks(x, labels)
ax.bar_label(rects1, labels=[f'{x:,.0f}' for x in rects1.datavalues], padding=1)
ax.bar_label(rects2, labels=[f'{x:,.0f}' for x in rects2.datavalues], padding=1)
ax.bar_label(rects3, labels=[f'{x:,.0f}' for x in rects3.datavalues], padding=1)
ax.bar_label(rects4, labels=[f'{x:,.0f}' for x in rects4.datavalues], padding=1)
ax.legend(fancybox=True)

fig.tight_layout()
plt.show()
