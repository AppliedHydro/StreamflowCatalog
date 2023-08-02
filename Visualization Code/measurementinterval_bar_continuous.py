import openpyxl as px             #handles catalog as dataframe
import matplotlib.pyplot as plt   #visualizes data
import numpy as np                #mathematical library with built-in calculation tools
import matplotlib as mpl

#import catalog as openpy dataframe iterable
workbook = px.open('C:/Users/sjsch/Desktop/Kendra/Streamflow_Catalog.xlsx')
wb = workbook.active
counter = 1
unique = ['None',]

hourly,daily,unknown = [],[],[]
min15,instant,annual = [],[],[]
min30, monthly = [],[]

#iterators to sort through unique values and evaluate measurement intervals
for row in wb.iter_rows(min_row=2,max_row=26883,min_col=20,max_col=20):
    for cell in row:
        if str(cell.value) in unique:
            pass
        elif str(cell.value) == 'None':
            pass
        else:
            unique.append(cell.value)
for rows in wb.iter_rows(min_row=2,max_row=26883,min_col=20,max_col=20):
    for cell in rows:
        counter += 1
        var = str(cell.value).lower()
        if str(wb.cell(counter,21).value).lower() == 'continuous': #subset by continuous
            if var == 'hourly':
                hourly.append(var)
            elif var == 'daily':
                daily.append(var)
            elif var == 'unknown':
                unknown.append(var)
            elif var == '30 minutes':
                min30.append(var)
            elif var == 'monthly':
                monthly.append(var)
            elif var == 'none':
                if str(wb.cell(counter,1).value) == 'None':
                    pass
                else:
                    unknown.append(var)
            elif var == '15 minutes':
                min15.append(var)
            elif var == 'instantaneous':
                instant.append(var)
            elif var == 'anually':
                annual.append(var)
            else:
                print(var)

#chart graphic
labels = ['instant','15 minutes','30 minutes','hourly','daily','monthly','annually','unknown']
counts = [len(instant),len(min15),len(min30),len(hourly),len(daily),len(monthly),len(annual),len(unknown)]
x = np.arange(len(labels))
width = 0.3
bar_colors = ['darkorange', 'darkgoldenrod', 'gold', 'orangered','lightsalmon','yellow','orange','red']
fig, ax = plt.subplots()
rects1 = ax.bar(x - width/8,counts,label='labels',color=bar_colors)
ax.yaxis.set_major_formatter(mpl.ticker.StrMethodFormatter('{x:,.0f}')) #format integers to #,### values
ax.set_ylabel('quantity')
ax.bar_label(rects1,labels=[f'{x:,.0f}' for x in rects1.datavalues],padding=0.5)
ax.set_xticks(x, labels,fontsize='x-small')

fig.tight_layout()
plt.show()
