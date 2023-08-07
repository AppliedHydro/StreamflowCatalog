import openpyxl as px
from datetime import datetime

def extract_date_components(date_value):
    if date_value is None:
        return None, None, None
    date_str = str(date_value)
    try:
        date_object = datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S')
        return date_object.year, date_object.month, date_object.day
    except ValueError:
        return None, None, None

file_path = 'C:/Users/sjsch/Desktop/Kendra/Streamflow_Catalog.xlsx' #Streamflow Catalog location on local machine
workbook = px.load_workbook(file_path)
worksheet = workbook.active

data_dict = {}

for row in worksheet.iter_rows(min_row=2, values_only=True):
    source = row[1]  
    
    # Only consider rows where the source is 'USGS' #Use the USGS organization tag as a filter
    if source != 'USGS':
        continue
    unique_identifier = row[2]  # Assuming the unique identifier is in the third column
    # Ignore rows with 'none', 'N/A', or blank unique identifiers
    if unique_identifier is None or (isinstance(unique_identifier, str) and unique_identifier.lower() in ('none', 'n/a', '')):
        continue
    # Ignore rows with non-numeric unique identifiers
    if not isinstance(unique_identifier, (int, float)):
        continue
    
    start_date, start_month, start_day = extract_date_components(row[14])      end_date, end_month, end_day = extract_date_components(row[15])  
    
    if unique_identifier in data_dict:
        data_dict[unique_identifier]['rows'].append(row)
        if start_date is not None and (data_dict[unique_identifier]['start_date'] is None or start_date < data_dict[unique_identifier]['start_date']):
            data_dict[unique_identifier]['start_date'] = start_date
            data_dict[unique_identifier]['start_month'] = start_month
            data_dict[unique_identifier]['start_day'] = start_day
        if end_date is not None and (data_dict[unique_identifier]['end_date'] is None or end_date > data_dict[unique_identifier]['end_date']):
            data_dict[unique_identifier]['end_date'] = end_date
            data_dict[unique_identifier]['end_month'] = end_month
            data_dict[unique_identifier]['end_day'] = end_day
    else:
        data_dict[unique_identifier] = {
            'rows': [row],
            'start_date': start_date,
            'start_month': start_month,
            'start_day': start_day,
            'end_date': end_date,
            'end_month': end_month,
            'end_day': end_day
        }

combined_workbook = px.Workbook()
combined_worksheet = combined_workbook.active

header_row = list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
combined_worksheet.append(header_row)

for unique_identifier, data in data_dict.items():
    start_date_str = f"{data['start_month']:02d}/{data['start_day']:02d}/{data['start_date']}" if data['start_date'] is not None else ''
    end_date_str = f"{data['end_month']:02d}/{data['end_day']:02d}/{data['end_date']}" if data['end_date'] is not None else ''
    
    combined_row = [unique_identifier] + list(data['rows'][0])[1:14] + [start_date_str, end_date_str] + list(data['rows'][0])[16:]
    combined_worksheet.append(combined_row)
    
#####
#To ensure no cross contamination with original streamflow catalog, the new values are written into an excel file which can then be added to original catalog
#####

combined_file_path = "C:/Users/sjsch/Desktop/Kendra/Streamflow_Edit_Duplicates.xlsx"
combined_workbook.save(combined_file_path)
