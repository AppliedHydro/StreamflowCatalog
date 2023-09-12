#####Editing_Master.py#####

#####################################################################################
#This is a master file with sections used for text/value editing in the catalog.    # 
#The scripts were run in discrete sections rather than as a top-to-bottom execution #.
#####################################################################################

#Variables
#w_dir (line 17) : local file path for streamflow catalog in .xlsx format

import openpyxl as px             #handles catalog as dataframe
import matplotlib.pyplot as plt   #visualizes data
import numpy as np                #mathematical library with built-in calculation tools
import matplotlib as mpl
from datetime import datetime, date
from collections import Counter

w_dir = 'C:/Users/sjsch/Desktop/Kendra/Streamflow_Catalog.xlsx' #streamflow catalog pathway
workbook = px.open(w_dir)
wb = workbook.active
maxrow = wb.max_row

#------------------------------------------------------
#checking for streamflow catalog length
#------------------------------------------------------
def get_length(wb): 
    for i in range(1, ws.max_row):
        if ws.cell(i, 1).value is None:
            index_row.append(i)
    for row_del in range(len(index_row)):
        ws.delete_rows(idx=index_row[row_del], amount=1)
    print(f"Streamflow catalog is {maxrow} lines")
    return maxrow

#------------------------------------------------------
#replacing dates with YYYY with 1/1/YYYY for consistant formatting
#------------------------------------------------------

def year_editor():
    err_msg = 'incorrect year format values'
    counter = 0
    for rows in wb.iter_rows(min_row=1,max_row=12671,min_col=1,max_col=1): #gage start dates
        for cell in rows:
            counter += 1
            date = str(cell.value)wb.cell(34406,1)
            assert len(date) == 4,err_msg
            cell.value = ('1/1/'+date)
    counter = 0
    for rows in wb.iter_rows(min_row=1,max_row=12671,min_col=2,max_col=2): #gage end dates
        for cell in rows:
            counter += 1
            date = str(cell.value)
            cell.value = ('1/1/'+date)
    workbook.save('C:/Users/sjsch/Desktop/Kendra/Streamflow_catalog.xlsx')

#------------------------------------------------------
#checking for empty latitude/longitude fields
#------------------------------------------------------
def empty_latlon():
    counter = 1 #skips header
    checks = [] #this list will have the row numbers of any empties
    for rows in wb.iter_rows(min_row=2,max_row=maxrow,min_col=11,max_col=11):
        for cell in rows:
            counter += 1
            if str(cell.value).lower() == 'none':
                checks.append(counter)
            else:
                pass

#------------------------------------------------------
#checking for empty date fields/erroneous default values
#------------------------------------------------------
def date_correct() -> list:
    years_start, years_end, unknown = [],[],[]
    counter = 1
    date_format = ('%Y-%m-%d %H:%M:%S')
    today = datetime.now()
    today = today.strftime(date_format)
    try:
        for row in wb.iter_rows(min_col=15,max_col=15,min_row=2,max_row=maxrow):
            for cell in row:
                counter += 1
                var = str(cell.value).lower()
                if var == 'none' or var == 'NA':         #skips blank rows                    
                    if str(wb.cell(counter,1).value) == 'None':
                        continue
                    else:
                        unknown.append(var)
                        continue
                elif var == 'unknown':
                    unknown.append(var)
                    continue
                elif var == '2/1/1857':         #autopopulated value found in some datasets
                    years_start.append('1857')
                    years_end.append('2022')
                    break
                elif str(wb.cell(counter,15).value) == '00:00:00':
                    continue
                else:
                    if str(wb.cell(counter,16).value) == 'None' or str(wb.cell(counter,16).value).lower() == 'unknown' or str(wb.cell(counter,16).value) == 'NA':
                        unknown.append(var)
                        continue
                if (wb.cell(counter,16).value) == 'current':
                    wb.cell(counter,16).value = today
                var_s, var_e = str(wb.cell(counter,15).value), str(wb.cell(counter,16).value)
                try:
                    date_s = datetime.strptime(var_s, date_format)
                except ValueError:
                    date_s = datetime.strptime(var_s, '%Y-%m-%d')
                try:
                    date_e = datetime.strptime(var_e, date_format)
                except ValueError:
                    date_e = datetime.strptime(var_e, '%Y-%m-%d')
                years_start.append(date_s.year)
                years_end.append(date_e.year)
                continue
        start = (Counter(years_start))
        end = (Counter(years_end))
        return start, end
    except:  
        print(f'Anomolous date parameter found at row {counter}')
        raise

#------------------------------------------------------
#checking for streamtype values subset by state, gage quantites by state
#------------------------------------------------------

import openpyxl as px
import operator as op
from collections import Counter

w_dir = 'C:/Users/sjsch/Desktop/Kendra/Streamflow_Catalog.xlsx' #streamflow catalog pathway
workbook = px.open(w_dir)
wb = workbook.active
maxrow = wb.max_row

#Creates a 'streamtype' class where state,stream types can be easily accessed
class streamtype:
    def __init__(self,stream,state):
        self.stream = stream
        self.state = state
        
def find_unique() -> list(unique):
    """
    iterates through streamtypes in streamflow catalog to create list of
    unique names. Can be use as reference
    """
    global unique
    unique = []   
    for row in wb.iter_rows(min_col=9,max_col=9,min_row=2,max_row=maxrow): #row 9 = streamtype
        for cell in row:
            cell.value = str(cell.value).rstrip() #remove problematic white spaces
            if str(cell.value) not in unique:
                unique.append(str(cell.value))
                continue
            elif str(cell.value) in unique:
                pass
    return unique

def make_list() -> list(type_data):
    """
    lines 113 - 126 must be run to execute successfully. Return object will be
    list variable of streamtype objects for stream_counter()
    """
    global type_data
    type_data = []
    counter = 1
    for row in wb.iter_rows(min_col=9,max_col=9,min_row=2,max_row=maxrow):
        for cell in row:
            type_data.append(streamtype(wb.cell(counter,9).value,wb.cell(counter,6).value))
            counter += 1
    type_data = list(filter(lambda item: item.state is not None, type_data))
    print('List name: type_data')

def stream_counter(type_data):
    """
    Using the stream_counter function, you can pass state variables to it and it will 
    parse number of streams of each type subset by state.
    @type_data : list of streamtype class objects with self.stream, self.state attributes
    """
    allowed = ['Idaho','Washington','Oregon']
    usr_state = input('input full state name (ex. Idaho): ')
    err_msg = 'State name not recognized. Check spelling/capitalization'
    assert usr_state in allowed,err_msg
    stream_total = []
    for val in type_data:
        if usr_state == val.state:
            stream_total.append(val)
    print('\n' * 3)
    print(f'{len(stream_total)} gages located in {usr_state}')
    strm = sum(val.stream == 'stream' for val in stream_total)
    canl = sum(val.stream == 'canal' for val in stream_total)
    weir = sum(val.stream == 'weir' for val in stream_total)
    print(f'There are: \na. {strm} streams\nb. {canl} canals\nc. {weir} weirs')
    print('\n' * 3)
    stream_counter(type_data)


def state_counter(type_data) -> int:
    usr_state = input('Input state name (ex. Washington): ')
    err_msg = 'State name not recognized. Check spelling/capitalization'
    allowed = ['Idaho','Washington','Oregon']
    assert usr_state in allowed,err_msg
    return sum(val.state == usr_state for val in type_data)
    
