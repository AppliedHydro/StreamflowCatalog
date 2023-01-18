from datetime import datetime
import openpyxl as px           #handles catalog data as dataframe object
import matplotlib.pyplot as plt #visualizes data 
import numpy as np              #collections of calculation tools

#import catalog as dataframe object
workbook = px.open('C:/Users/sjsch/Desktop/Kendra/Streamflow_Catalog.xlsx')
wb = workbook.active

years_start = []
years_end = []
unknown = []
counter = 1

#iterate through date values, discard 'unknowns' and convert rest into datetime object
try:
    for row in wb.iter_rows(min_col=15,max_col=15,min_row=2,max_row=26871):
        for cell in row:
            counter += 1
            date_format = ('%Y-%m-%d %H:%M:%S')
            date_format1 = ('%H:%M:%S')
            var = str(cell.value).lower()
            if str(wb.cell(counter,16).value) == 'current':
                cell.value = datetime.now()
                pass
            elif str(wb.cell(counter,1).value).lower() == 'none':
                continue
            elif str(wb.cell(counter,21).value).lower() == 'continuous': #filter by continuous
                if var == 'none':
                    if str(wb.cell(counter,16).value) == 'None':
                        continue
                    else:
                        unknown.append(var)
                        continue
                elif var == 'unknown':
                    unknown.append(var)
                    continue
                elif var == '2/1/1857':
                    years_start.append('1857')
                    years_end.append('2022')
                elif str(wb.cell(counter,15).value) == '00:00:00':
                    continue
                else:
                    if str(wb.cell(counter,16).value) == 'None' or str(wb.cell(counter,16).value).lower() == 'unknown':
                        unknown.append(var)
                        continue
                    date_s = datetime.strptime(var,date_format) #pulls date from excel string as datetime object
                    years_start.append(date_s.year)
                    date_e = datetime.strptime(str(wb.cell(counter,16).value),date_format)
                    years_end.append(date_e.year)
                    continue
            else:
                pass
except: #checks to ensure all values have been handled
    print(counter)
    raise

totals = []
counter = 0
try:        #calculate time total between start and end dates
    for i in range(0,25526):
        var = int(years_end[counter]) - int(years_start[counter])
        totals.append(var)
        counter += 1
except:
    pass

#chart graphic
dist1 = totals
dist2 = range(0,126,10)
fig, axs = plt.subplots(tight_layout=True)
axs.hist(dist1,bins=125,range=(0,125),color='orange')
axs.set_xticks(dist2)
axs.set_ylabel('quantity')
axs.set_xlabel('years of operation')
axs.set_xlim([0,100]) #constrain x-axis limitations for viewability
plt.show()
